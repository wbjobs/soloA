import os
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from django.conf import settings
from .analytics_service import AnalyticsService


class ExportService:
    def __init__(self):
        self.analytics = AnalyticsService()
        self._setup_chinese_font()

    def _setup_chinese_font(self):
        try:
            font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'SimHei.ttf')
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('SimHei', font_path))
                plt.rcParams['font.sans-serif'] = ['SimHei']
        except:
            plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False

    def generate_pdf_report(self, report, start_date, end_date):
        filename = f'report_{report.report_type}_{start_date}_{end_date}.pdf'
        filepath = os.path.join(settings.EXPORT_DIR, filename)
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            alignment=1,
            spaceAfter=20
        )
        
        story = []
        
        titles = {
            'overview': '电商数据概览报表',
            'user': '用户行为分析报表',
            'product': '商品销售分析报表',
        }
        
        story.append(Paragraph(titles.get(report.report_type, '分析报表'), title_style))
        story.append(Paragraph(f'统计周期: {start_date} 至 {end_date}', styles['Normal']))
        story.append(Spacer(1, 12))
        
        if report.report_type == 'overview':
            story.extend(self._generate_overview_content(start_date, end_date, styles))
        elif report.report_type == 'user':
            story.extend(self._generate_user_content(start_date, end_date, styles))
        elif report.report_type == 'product':
            story.extend(self._generate_product_content(start_date, end_date, styles))
        
        story.append(Spacer(1, 20))
        story.append(Paragraph(f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Italic']))
        
        doc.build(story)
        
        return filename

    def _generate_overview_content(self, start_date, end_date, styles):
        story = []
        
        stats = self.analytics.get_overview_stats(start_date, end_date)
        
        story.append(Paragraph('一、核心指标', styles['Heading2']))
        
        metrics_data = [
            ['指标', '数值'],
            ['页面浏览量(PV)', str(stats['pv'])],
            ['独立访客数(UV)', str(stats['uv'])],
            ['点击量', str(stats['clicks'])],
            ['加购数', str(stats['add_to_carts'])],
            ['订单数', str(stats['purchases'])],
            ['总收入', f"¥{stats['total_revenue']}"],
            ['点击率', f"{stats['click_through_rate']}%"],
            ['整体转化率', f"{stats['overall_conversion_rate']}%"],
            ['客单价', f"¥{stats['avg_order_value']}"],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[2.5*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph('二、趋势图表', styles['Heading2']))
        
        trend = self.analytics.get_daily_trend(start_date, end_date)
        chart_path = self._generate_trend_chart(trend)
        if chart_path:
            story.append(Image(chart_path, width=8*inch, height=4*inch))
            story.append(Spacer(1, 10))
        
        return story

    def _generate_user_content(self, start_date, end_date, styles):
        story = []
        
        story.append(Paragraph('一、转化漏斗', styles['Heading2']))
        
        funnel = self.analytics.get_conversion_funnel(start_date, end_date)
        funnel_data = [['阶段', '用户数', '转化率', '流失率']]
        for stage in funnel:
            funnel_data.append([
                stage['name'],
                str(stage['users']),
                f"{stage['percentage']}%",
                f"{stage['drop_off']}%"
            ])
        
        funnel_table = Table(funnel_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        funnel_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(funnel_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph('二、用户分群', styles['Heading2']))
        
        segments = self.analytics.get_user_segment_distribution()
        segment_data = [['用户分群', '人数']]
        for seg in segments:
            segment_data.append([seg['name'], str(seg['value'])])
        
        segment_table = Table(segment_data, colWidths=[2*inch, 2*inch])
        segment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(segment_table)
        story.append(Spacer(1, 15))
        
        repeat = self.analytics.get_repeat_purchase_rate(start_date, end_date)
        story.append(Paragraph(f"复购率: {repeat['repeat_purchase_rate']}%", styles['Normal']))
        story.append(Paragraph(f"总购买用户: {repeat['total_buyers']}人", styles['Normal']))
        story.append(Paragraph(f"复购用户: {repeat['repeat_buyers']}人", styles['Normal']))
        
        return story

    def _generate_product_content(self, start_date, end_date, styles):
        story = []
        
        story.append(Paragraph('一、商品排行榜', styles['Heading2']))
        
        products = self.analytics.get_product_performance(start_date, end_date, limit=10)
        
        product_data = [['商品ID', '浏览', '点击', '加购', '下单', '收入', '点击率', '转化率']]
        for p in products:
            product_data.append([
                p['product_id'],
                str(p['views']),
                str(p['clicks']),
                str(p['add_to_carts']),
                str(p['purchases']),
                f"¥{p['revenue']}",
                f"{p['click_through_rate']}%",
                f"{p['conversion_rate']}%",
            ])
        
        col_widths = [1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch, 0.8*inch, 0.8*inch]
        product_table = Table(product_data, colWidths=col_widths)
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(product_table)
        
        return story

    def _generate_trend_chart(self, trend):
        if not trend['dates']:
            return None
        
        fig, ax1 = plt.subplots(figsize=(10, 5))
        
        x = range(len(trend['dates']))
        
        ax1.plot(x, trend['pv'], 'b-o', label='PV', linewidth=2)
        ax1.plot(x, trend['uv'], 'g-s', label='UV', linewidth=2)
        ax1.set_xlabel('日期')
        ax1.set_ylabel('访问量')
        ax1.legend(loc='upper left')
        ax1.set_xticks(x)
        ax1.set_xticklabels([d[-5:] for d in trend['dates']], rotation=45)
        ax1.grid(True, alpha=0.3)
        
        ax2 = ax1.twinx()
        ax2.bar(x, trend['orders'], alpha=0.3, color='r', label='订单数')
        ax2.set_ylabel('订单数')
        ax2.legend(loc='upper right')
        
        plt.title('用户行为趋势')
        plt.tight_layout()
        
        chart_path = os.path.join(settings.EXPORT_DIR, 'trend_chart.png')
        plt.savefig(chart_path, dpi=100)
        plt.close()
        
        return chart_path

    def generate_excel_report(self, report, start_date, end_date):
        filename = f'report_{report.report_type}_{start_date}_{end_date}.xlsx'
        filepath = os.path.join(settings.EXPORT_DIR, filename)
        
        wb = openpyxl.Workbook()
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def apply_style(sheet, headers):
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
        
        if report.report_type == 'overview':
            self._fill_overview_sheet(wb, start_date, end_date, apply_style, center_align, thin_border)
        elif report.report_type == 'user':
            self._fill_user_sheet(wb, start_date, end_date, apply_style, center_align, thin_border)
        elif report.report_type == 'product':
            self._fill_product_sheet(wb, start_date, end_date, apply_style, center_align, thin_border)
        
        wb.save(filepath)
        return filename

    def _fill_overview_sheet(self, wb, start_date, end_date, apply_style, center_align, thin_border):
        stats = self.analytics.get_overview_stats(start_date, end_date)
        
        ws = wb.active
        ws.title = '核心指标'
        
        apply_style(ws, ['指标', '数值'])
        
        data = [
            ['页面浏览量(PV)', stats['pv']],
            ['独立访客数(UV)', stats['uv']],
            ['点击量', stats['clicks']],
            ['加购数', stats['add_to_carts']],
            ['订单数', stats['purchases']],
            ['总收入(元)', stats['total_revenue']],
            ['点击率(%)', stats['click_through_rate']],
            ['整体转化率(%)', stats['overall_conversion_rate']],
            ['客单价(元)', stats['avg_order_value']],
        ]
        
        for row_idx, row in enumerate(data, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_align
                cell.border = thin_border
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        ws_trend = wb.create_sheet('趋势数据')
        trend = self.analytics.get_daily_trend(start_date, end_date)
        
        apply_style(ws_trend, ['日期', 'PV', 'UV', '订单数', '收入(元)'])
        
        for row_idx, date in enumerate(trend['dates'], 2):
            ws_trend.cell(row=row_idx, column=1, value=date).border = thin_border
            ws_trend.cell(row=row_idx, column=2, value=trend['pv'][row_idx-2] if row_idx-2 < len(trend['pv']) else 0).border = thin_border
            ws_trend.cell(row=row_idx, column=3, value=trend['uv'][row_idx-2] if row_idx-2 < len(trend['uv']) else 0).border = thin_border
            ws_trend.cell(row=row_idx, column=4, value=trend['orders'][row_idx-2] if row_idx-2 < len(trend['orders']) else 0).border = thin_border
            ws_trend.cell(row=row_idx, column=5, value=trend['revenue'][row_idx-2] if row_idx-2 < len(trend['revenue']) else 0).border = thin_border

    def _fill_user_sheet(self, wb, start_date, end_date, apply_style, center_align, thin_border):
        ws = wb.active
        ws.title = '转化漏斗'
        
        funnel = self.analytics.get_conversion_funnel(start_date, end_date)
        apply_style(ws, ['阶段', '用户数', '转化率(%)', '流失率(%)'])
        
        for row_idx, stage in enumerate(funnel, 2):
            ws.cell(row=row_idx, column=1, value=stage['name']).border = thin_border
            ws.cell(row=row_idx, column=2, value=stage['users']).border = thin_border
            ws.cell(row=row_idx, column=3, value=stage['percentage']).border = thin_border
            ws.cell(row=row_idx, column=4, value=stage['drop_off']).border = thin_border
        
        ws_seg = wb.create_sheet('用户分群')
        segments = self.analytics.get_user_segment_distribution()
        apply_style(ws_seg, ['用户分群', '人数'])
        
        for row_idx, seg in enumerate(segments, 2):
            ws_seg.cell(row=row_idx, column=1, value=seg['name']).border = thin_border
            ws_seg.cell(row=row_idx, column=2, value=seg['value']).border = thin_border
        
        ws_repeat = wb.create_sheet('复购分析')
        repeat = self.analytics.get_repeat_purchase_rate(start_date, end_date)
        apply_style(ws_repeat, ['指标', '数值'])
        repeat_data = [
            ['总购买用户', repeat['total_buyers']],
            ['复购用户', repeat['repeat_buyers']],
            ['复购率(%)', repeat['repeat_purchase_rate']],
        ]
        for row_idx, row in enumerate(repeat_data, 2):
            for col_idx, value in enumerate(row, 1):
                ws_repeat.cell(row=row_idx, column=col_idx, value=value).border = thin_border

    def _fill_product_sheet(self, wb, start_date, end_date, apply_style, center_align, thin_border):
        ws = wb.active
        ws.title = '商品排行'
        
        products = self.analytics.get_product_performance(start_date, end_date, limit=50)
        apply_style(ws, ['商品ID', '浏览', '点击', '加购', '下单', '收入(元)', '点击率(%)', '转化率(%)'])
        
        for row_idx, p in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=p['product_id']).border = thin_border
            ws.cell(row=row_idx, column=2, value=p['views']).border = thin_border
            ws.cell(row=row_idx, column=3, value=p['clicks']).border = thin_border
            ws.cell(row=row_idx, column=4, value=p['add_to_carts']).border = thin_border
            ws.cell(row=row_idx, column=5, value=p['purchases']).border = thin_border
            ws.cell(row=row_idx, column=6, value=p['revenue']).border = thin_border
            ws.cell(row=row_idx, column=7, value=p['click_through_rate']).border = thin_border
            ws.cell(row=row_idx, column=8, value=p['conversion_rate']).border = thin_border
